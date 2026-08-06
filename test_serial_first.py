import io
import os
import tempfile
import unittest
from pathlib import Path
from unittest import mock

os.environ.setdefault("ZTP_DEV", "1")
import app


VALID_CONFIG = "system { root-authentication { encrypted-password x; } }\n"


class SerialFirstWorkflowTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        root = Path(self.tmp.name)
        names = ("DATA_DIR", "NGINX_DIR", "UPLOAD_DIR", "DEVICES_JSON", "STATIC_MAPPINGS_JSON",
                 "PROFILES_JSON", "SETTINGS_JSON", "CREDS_JSON", "PROVISIONING_STATE_JSON",
                 "CONFIG_POOL_JSON", "ASSIGNMENTS_JSON", "RESULTS_JSON", "HISTORY_JSONL",
                 "DEVICE_RUNTIME_JSON", "DOWNLOAD_RECORDS_JSON", "PARSER_CURSORS_JSON",
                 "ALLOCATION_LOCK", "HISTORY_LOCK", "LEASES_FILE", "SYSLOG_FILE", "NGINX_ACCESS",
                 "DHCPD_CONF", "DEV_MODE")
        self.old = {name: getattr(app, name) for name in names}
        app.DATA_DIR = root / "state"
        app.NGINX_DIR = root / "configs"
        app.UPLOAD_DIR = app.NGINX_DIR
        app.DATA_DIR.mkdir()
        app.NGINX_DIR.mkdir()
        for name in ("DEVICES_JSON", "STATIC_MAPPINGS_JSON", "PROFILES_JSON", "SETTINGS_JSON",
                     "CREDS_JSON", "PROVISIONING_STATE_JSON", "CONFIG_POOL_JSON", "ASSIGNMENTS_JSON",
                     "RESULTS_JSON", "HISTORY_JSONL", "DEVICE_RUNTIME_JSON", "DOWNLOAD_RECORDS_JSON",
                     "PARSER_CURSORS_JSON"):
            setattr(app, name, app.DATA_DIR / Path(getattr(app, name)).name)
        app.ALLOCATION_LOCK = app.DATA_DIR / ".allocation.lock"
        app.HISTORY_LOCK = app.DATA_DIR / ".history.lock"
        app.LEASES_FILE = root / "dhcpd.leases"
        app.SYSLOG_FILE = root / "syslog"
        app.NGINX_ACCESS = root / "access.log"
        app.DHCPD_CONF = root / "dhcpd.conf"
        app.DEV_MODE = True
        app.app.config.update(TESTING=True)
        settings = dict(app.DEFAULT_SETTINGS)
        settings.update({"active_mode": "ZTP_PROVISIONING", "operating_mode": "ZTP_PROVISIONING",
                         "global_mode": "ZTP_PROVISIONING", "pending_mode": "",
                         "project_expected_devices": "3"})
        app.write_settings(settings)
        state = app._default_provisioning_state()
        state["project"].update({"status": "ACTIVE", "expected_devices": 3, "next_sequence": 1})
        app.commit_provisioning_state(state)
        app.write_devices([])
        app.write_profiles([])

    def tearDown(self):
        for name, value in self.old.items():
            setattr(app, name, value)
        self.tmp.cleanup()

    def add_profile(self, prefix="Juniper-ex4100-h-12mp-", model="EX4100-H-12MP",
                    pool="H12", pattern="*"):
        rows = app.read_profiles()
        rows.append({"label": model, "vendor_class": prefix, "vendor_prefix": prefix,
                     "device_model": model, "match_mode": "contains", "config_file": "",
                     "assignment_type": "AUTO", "pool_name": pool,
                     "config_pattern": pattern,
                     "compatibility_group": model, "option60_confirmed": "yes"})
        app.write_profiles(rows)

    def add_config(self, filename, model="EX4100-H-12MP", pool="H12", order=1):
        path = app.NGINX_DIR / filename
        path.write_text(VALID_CONFIG, encoding="utf-8")
        state = app.read_provisioning_state()
        state["configs"][filename] = {
            "status": "AVAILABLE", "checksum": app.config_sha256(path),
            "allocation_order": order, "assigned_device": "", "assigned_serial": "",
            "file_size": path.stat().st_size, "device_model": model,
            "supported_models": [model], "pool_name": pool, "auto_pool_enabled": True,
        }
        app.commit_provisioning_state(state)

    @staticmethod
    def lease(serial="GE4825AW015", prefix="Juniper-ex4100-h-12mp-",
              mac="aa:bb:cc:dd:ee:01"):
        return {"mac": mac, "client_id": "", "hostname": serial, "state": "active",
                "option60": prefix + serial}

    def reserve(self, serial="GE4825AW015", mac="aa:bb:cc:dd:ee:01", ip="192.168.250.20",
                prefix="Juniper-ex4100-h-12mp-"):
        return app.reserve_project_assignment(ip, self.lease(serial, prefix, mac))

    def test_parse_supported_option60_models(self):
        profiles = []
        for prefix, (model, pool) in app.KNOWN_VENDOR_PREFIXES.items():
            profiles.append({"label": model, "vendor_prefix": prefix, "vendor_class": prefix,
                             "device_model": model, "pool_name": pool, "match_mode": "contains"})
        for prefix, (model, _) in app.KNOWN_VENDOR_PREFIXES.items():
            identity, error = app.parse_option60_identity(prefix + "GE4825AW015", profiles)
            self.assertEqual("", error)
            self.assertEqual("GE4825AW015", identity["serial"])
            self.assertEqual(model, identity["device_model"])

    def test_reject_missing_or_malformed_option60(self):
        self.assertEqual("OPTION60_NOT_CAPTURED", app.parse_option60_identity("")[1])
        self.assertEqual("SERIAL_NOT_PARSED", app.parse_option60_identity("Juniper-ex4100-h-12mp-")[1])
        self.assertEqual("SERIAL_NOT_PARSED", app.parse_option60_identity("bad value")[1])

    def test_generated_dhcp_and_lease_capture_option60(self):
        self.assertIn("set vendor-string = option vendor-class-identifier;", app.generate_dhcpd())
        app.LEASES_FILE.write_text(
            'lease 192.168.250.20 { hardware ethernet aa:bb:cc:dd:ee:01; '
            'binding state active; set vendor-string = "Juniper-ex4100-h-12mp-GE4825AW015"; }\n',
            encoding="utf-8")
        lease = app.parse_leases()["192.168.250.20"]
        self.assertEqual("GE4825AW015", lease["serial"])
        self.assertEqual("Juniper-ex4100-h-12mp-GE4825AW015", lease["option60"])

    def test_profile_selects_exact_named_pool_without_global_fallback(self):
        self.add_profile()
        self.add_config("wrong.conf", pool="OTHER", order=1)
        self.add_config("right.conf", pool="H12", order=2)
        assignment, error = self.reserve()
        self.assertEqual("", error)
        self.assertEqual("right.conf", assignment["filename"])
        self.assertEqual("serial:ge4825aw015", assignment["device_key"])

    def test_profile_filename_pattern_is_enforced_inside_named_pool(self):
        self.add_profile(pattern="OXISANTA_EX4100_PC*")
        self.add_config("OTHER.conf", pool="H12", order=1)
        self.add_config("OXISANTA_EX4100_PC001.conf.txt", pool="H12", order=2)
        assignment, error = self.reserve()
        self.assertEqual("", error)
        self.assertEqual("OXISANTA_EX4100_PC001.conf.txt", assignment["filename"])
        self.assertEqual("OXISANTA_EX4100_PC*", assignment["config_pattern"])

    def test_profile_filename_pattern_without_matching_config_fails_closed(self):
        self.add_profile(pattern="OXISANTA_EX4100_PC*")
        self.add_config("OTHER.conf", pool="H12")
        assignment, error = self.reserve()
        self.assertIsNone(assignment)
        self.assertEqual("PROFILE_PATTERN_EMPTY", error)

    def test_no_profile_and_ambiguous_profile_fail_closed(self):
        self.add_config("one.conf")
        self.assertEqual("PROFILE_NOT_MATCHED", self.reserve()[1])
        self.add_profile()
        self.add_profile()
        self.assertEqual("AMBIGUOUS_PROFILE", self.reserve()[1])

    def test_override_wins_over_profile(self):
        self.add_profile(pool="PROFILE")
        self.add_config("profile.conf", pool="PROFILE")
        self.add_config("override.conf", pool="OVERRIDE")
        app.write_devices([{"match_method": "serial", "serial_number": "GE4825AW015",
                            "mac_address": "", "device_type": "EX4100-H-12MP",
                            "expected_model": "EX4100-H-12MP", "hostname": "GE4825AW015",
                            "ip_address": "", "mgmt_ip": "", "client_id": "",
                            "compatibility_group": "", "specific_config_file": "override.conf",
                            "assignment_type": "STATIC", "pool_name": "",
                            "option60_confirmed": "yes"}])
        assignment, error = self.reserve()
        self.assertEqual("", error)
        self.assertEqual("override.conf", assignment["filename"])
        self.assertEqual("Override", assignment["assignment_source"])

    def test_model_isolation_and_override_cannot_bypass(self):
        self.add_profile(prefix="Juniper-ex4100-24p-", model="EX4100-24P", pool="P24")
        self.add_config("24t.conf", model="EX4100-24T", pool="P24")
        self.assertEqual("MODEL_CONFIG_MISMATCH",
                         self.reserve(prefix="Juniper-ex4100-24p-")[1])
        app.write_devices([{"match_method": "serial", "serial_number": "GE4825AW015",
                            "mac_address": "", "device_type": "EX4100-24P", "expected_model": "EX4100-24P",
                            "hostname": "GE4825AW015", "ip_address": "", "mgmt_ip": "", "client_id": "",
                            "compatibility_group": "", "specific_config_file": "24t.conf",
                            "assignment_type": "STATIC", "pool_name": "", "option60_confirmed": "yes"}])
        self.assertEqual("MODEL_CONFIG_MISMATCH",
                         self.reserve(prefix="Juniper-ex4100-24p-")[1])

    def test_same_serial_new_mac_reuses_config_and_mac_never_allocates(self):
        self.add_profile(); self.add_config("one.conf"); self.add_config("two.conf", order=2)
        first, _ = self.reserve()
        second, _ = self.reserve(mac="aa:bb:cc:dd:ee:99", ip="192.168.250.21")
        self.assertEqual(first["filename"], second["filename"])
        self.assertEqual("aa:bb:cc:dd:ee:99", second["observed_mac"])
        third, _ = self.reserve(serial="GE4825AW016", mac="aa:bb:cc:dd:ee:99",
                                ip="192.168.250.22")
        self.assertNotEqual(first["device_key"], third["device_key"])
        self.assertNotEqual(first["filename"], third["filename"])

    def test_retry_and_ownership_are_serial_safe(self):
        self.add_profile(); self.add_config("one.conf")
        first, _ = self.reserve()
        retry, _ = self.reserve()
        self.assertEqual(first["filename"], retry["filename"])
        other, error = self.reserve(serial="GE4825AW016", ip="192.168.250.21")
        self.assertIsNone(other)
        self.assertEqual("PROFILE_POOL_EMPTY", error)
        state = app.read_provisioning_state()
        self.assertEqual("GE4825AW015", state["configs"]["one.conf"]["assigned_serial"])

    def test_manual_download_by_other_serial_is_blocked(self):
        self.add_profile(); self.add_config("one.conf")
        self.reserve()
        app.LEASES_FILE.write_text(
            'lease 192.168.250.21 { hardware ethernet aa:bb:cc:dd:ee:02; binding state active; '
            'set vendor-string = "Juniper-ex4100-h-12mp-GE4825AW016"; }\n', encoding="utf-8")
        response = app.app.test_client().get("/ztp/config/one.conf",
                                             environ_base={"REMOTE_ADDR": "192.168.250.21"})
        self.assertEqual(409, response.status_code)
        self.assertIn(b"CONFIG_OWNERSHIP_CONFLICT", response.data)

    def test_release_completed_returns_config_and_preserves_rules_history(self):
        self.add_profile(); self.add_config("one.conf")
        assignment, _ = self.reserve()
        state = app.read_provisioning_state()
        state["devices"][assignment["device_key"]]["state"] = "DELIVERED"
        state["configs"]["one.conf"]["status"] = "DELIVERED"
        app.commit_provisioning_state(state)
        before_profiles = app.read_profiles()
        ok, _ = app.release_serial_assignment("GE4825AW015")
        self.assertTrue(ok)
        state = app.read_provisioning_state()
        self.assertNotIn("serial:ge4825aw015", state["devices"])
        self.assertEqual("AVAILABLE", state["configs"]["one.conf"]["status"])
        self.assertEqual(before_profiles, app.read_profiles())
        self.assertTrue(any(row["event_type"] == "RELEASE_ASSIGNMENT" for row in app.read_history()))

    def test_release_in_progress_returns_config_and_keeps_override(self):
        self.add_profile(); self.add_config("one.conf")
        app.write_devices([{"match_method": "serial", "serial_number": "GE4825AW015",
                            "mac_address": "", "device_type": "EX4100-H-12MP",
                            "expected_model": "EX4100-H-12MP", "hostname": "GE4825AW015",
                            "ip_address": "", "mgmt_ip": "", "client_id": "",
                            "compatibility_group": "", "specific_config_file": "one.conf",
                            "assignment_type": "STATIC", "pool_name": "",
                            "option60_confirmed": "yes"}])
        self.reserve()
        overrides = app.read_devices()
        ok, _ = app.release_serial_assignment("GE4825AW015")
        self.assertTrue(ok)
        self.assertEqual(overrides, app.read_devices())
        self.assertEqual("AVAILABLE", app.read_config_pool()[0]["status"])

    def test_three_results_report_and_legacy_verified(self):
        self.assertEqual("In Progress", app.ui_result("FETCHING"))
        self.assertEqual("Completed", app.ui_result("VERIFIED"))
        self.assertEqual("Error", app.ui_result("PROFILE_NOT_MATCHED"))
        response = app.app.test_client().get("/export/deployment-report.xlsx")
        self.assertEqual(200, response.status_code)
        self.assertTrue(response.data.startswith(b"PK"))

    def test_report_contains_serial_config_result_and_two_sheets(self):
        self.add_profile(); self.add_config("one.conf")
        assignment, _ = self.reserve()
        state = app.read_provisioning_state()
        state["devices"][assignment["device_key"]]["state"] = "DELIVERED"
        app.commit_provisioning_state(state)
        response = app.app.test_client().get("/export/deployment-report.xlsx")
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(response.data), read_only=True)
        self.assertEqual(["Summary", "Devices"], workbook.sheetnames)
        values = list(workbook["Devices"].values)
        self.assertEqual(("Serial Number", "Model", "Config File", "Result", "Last Update"),
                         values[0])
        self.assertEqual(("GE4825AW015", "EX4100-H-12MP", "one.conf", "Completed"),
                         values[1][:4])

    def test_reset_retest_preserves_configs_profiles_settings_and_history(self):
        self.add_profile(); self.add_config("one.conf"); self.reserve()
        app.append_history("TEST_KEEP")
        ok, _ = app.reset_workspace("RETEST")
        self.assertTrue(ok)
        self.assertTrue((app.NGINX_DIR / "one.conf").exists())
        self.assertTrue(app.read_profiles())
        self.assertEqual({}, app.read_assignments())
        self.assertEqual("AVAILABLE", app.read_config_pool()[0]["status"])
        self.assertTrue(any(row["event_type"] == "TEST_KEEP" for row in app.read_history()))

    def test_clean_reset_preserves_network_settings_and_history(self):
        self.add_profile(); self.add_config("one.conf"); self.reserve()
        settings = app.read_settings()
        settings.update({"server_ip": "192.168.250.1", "subnet": "192.168.250.0"})
        app.write_settings(settings)
        app.append_history("TEST_KEEP")
        ok, _ = app.reset_workspace("CLEAN")
        self.assertTrue(ok)
        self.assertEqual("192.168.250.1", app.read_settings()["server_ip"])
        self.assertEqual("192.168.250.0", app.read_settings()["subnet"])
        self.assertEqual([], app.read_profiles())
        self.assertEqual([], app.list_configs())
        self.assertTrue(any(row["event_type"] == "TEST_KEEP" for row in app.read_history()))

    def test_reset_rolls_back_if_dhcp_restart_fails(self):
        self.add_profile(); self.add_config("one.conf"); self.reserve()
        original_state = app.read_provisioning_state()
        calls = []

        def systemctl(action, service="isc-dhcp-server"):
            calls.append(action)
            if action == "start" and calls.count("start") == 1:
                return False, "simulated restart failure"
            return True, ""

        with mock.patch.object(app, "_systemctl_action", side_effect=systemctl):
            ok, message = app.reset_workspace("RETEST")
        self.assertFalse(ok)
        self.assertIn("previous state restored", message)
        restored = app.read_provisioning_state()
        self.assertEqual(original_state["devices"], restored["devices"])
        self.assertEqual(original_state["configs"], restored["configs"])
        self.assertTrue((app.NGINX_DIR / "one.conf").exists())

    def test_ui_is_serial_first_without_verify_or_mac_matching(self):
        response = app.app.test_client().get("/?view=mappings")
        text = response.get_data(as_text=True)
        self.assertIn("Device Override", text)
        self.assertIn("Test Option 60", text)
        self.assertNotIn("By MAC", text)
        self.assertNotIn(">Verify<", text)
        settings = app.app.test_client().get("/?view=settings").get_data(as_text=True)
        self.assertNotIn("Test Option 60", settings)

    def test_ui_shows_mapping_pattern_inventory_and_edit_action(self):
        self.add_profile(pattern="OXISANTA_EX4100_PC*")
        self.add_config("OXISANTA_EX4100_PC001.conf.txt")
        response = app.app.test_client().get("/?view=mappings")
        text = response.get_data(as_text=True)
        self.assertIn("Saved Vendor Mappings", text)
        self.assertIn("OXISANTA_EX4100_PC*", text)
        self.assertIn("1 matched", text)
        self.assertIn('/profiles/0/edit', text)
        self.assertIn("Save Changes", text)

    def test_saved_profile_can_be_edited_outside_ztp_mode(self):
        self.add_profile(pattern="OLD*")
        settings = app.read_settings()
        settings.update({"active_mode": "DHCP_FILE_SERVER",
                         "operating_mode": "DHCP_FILE_SERVER",
                         "global_mode": "DHCP_FILE_SERVER"})
        app.write_settings(settings)
        with mock.patch.object(app, "deploy_dhcpd", return_value=(True, "candidate valid")):
            response = app.app.test_client().post("/profiles/0/edit", data={
                "label": "EX4100-H-12MP",
                "vendor_prefix": "Juniper-ex4100-h-12mp-",
                "device_model": "EX4100-H-12MP",
                "pool_name": "OXISANTA_EX4100_H_12MP",
                "config_pattern": "OXISANTA_EX4100_PC*",
            })
        self.assertEqual(302, response.status_code)
        profile = app.read_profiles()[0]
        self.assertEqual("OXISANTA_EX4100_PC*", profile["config_pattern"])
        self.assertEqual("OXISANTA_EX4100_H_12MP", profile["pool_name"])

    def test_legacy_profile_pattern_migration_is_backed_up_and_idempotent(self):
        legacy = [{"label": "OXISANTA_EX4100", "vendor_class": "Juniper-ex4100-h-12mp-",
                   "vendor_prefix": "Juniper-ex4100-h-12mp-", "device_model": "EX4100-H-12MP",
                   "match_mode": "contains", "config_file": "", "assignment_type": "AUTO",
                   "pool_name": "OXISANTA_EX4100_H_12MP", "compatibility_group": "",
                   "option60_confirmed": "yes"}]
        app._atomic_write_json(app.PROFILES_JSON, legacy)
        self.assertTrue(app.migrate_profile_patterns())
        self.assertFalse(app.migrate_profile_patterns())
        self.assertEqual("OXISANTA_EX4100_PC*", app.read_profiles()[0]["config_pattern"])
        self.assertTrue((app.DATA_DIR / "migration-backup-profile-pattern-v27.0.1" /
                         "generic_profiles.json").exists())


if __name__ == "__main__":
    unittest.main()
